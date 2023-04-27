import os
import re
import openpyxl
import datetime
from colorama import Fore, Style
import tkinter as tk
from tkinter import messagebox


def onGenerateListButton():
    order_number = orderEntry.get()
    if not order_number.startswith('picking slip for '):
        shipping_list = shippingListGenerator(order_number)
        shippingListLabel.config(text=shipping_list) # Update the label text with the shipping list


def shippingListGenerator(order_num):
    try:
        file = find_file()
        print(file)

    except Exception as e:
        print(e)
        return 'Cannot find the production schedule Excel file'

    try:
        petco_order = False
        order = search_orders(order_num, file)

        if len(order['table']) == 0:
            order = search_in_petco_folder(order_num)
            if len(order['table']) != 0:
                petco_order = True

        if len(order['table']) == 0:
            return 'order no found'
        else:
            return generateShippingList(order, petco_order)

    except Exception as e:
        print(e)
        return 'Error while reading production schedule, please try manually'

    return f"Shipping list for order {order_number} is created."


def find_file():
    # Set the path to search in
    path = "../../../Prodution Schedule"

    # Get a list of all the files in the path
    files = os.listdir(path)

    # Initialize variables to keep track of the latest file and its modification date
    latest_file = None
    latest_date = None

    # Loop through the files and look for a file that starts with "AMERICAN JERKY ORDER"
    for file in files:
        if file.startswith("AMERICAN JERKY ORDER") and file.endswith(".xlsx"):
            # Get the modification date of the file
            file_path = os.path.join(path, file)
            file_date = os.path.getmtime(file_path)

            # If this is the first file or if this file has a later modification date, update the latest file and date
            if latest_date is None or file_date > latest_date:
                latest_file = file_path
                latest_date = file_date

    if latest_file is not None:
        # If we found the latest file, print its name and return the file path
        print("Found latest file:", latest_file)
        return latest_file
    else:
        # If we didn't find the file, print an error message and return None
        print("Error: file not found")
        return None


def search_orders(orderNumber, file):
    orderNumber = orderNumber.lower().replace(" ", "")

    # Open the file as a workbook
    workbook = openpyxl.load_workbook(file, data_only = True)

    # Set the sheet names to search
    # sheet_names = ["PETCO ", "WPP", "HEB", "SMP", "CANADA", "UPG"]
    sheet_names = ["UPG", "PETCO ", "WPP", "HEB", "SMP", "CANADA"]
    # sheet_names = ["UPG", "WPP", "HEB", "SMP", "CANADA"]

    customer = ''
    load_date = None
    ajc_order_number = ''
    ctm_order_number = ''
    output_table = []

    # Loop through the sheets
    for sheet_name in sheet_names:
        # Get the sheet by name
        sheet = workbook[sheet_name]

        # Set up variables to track the number of empty rows
        empty_row_count = 0
        last_row = sheet.max_row

        # Loop through the rows
        for row in reversed(range(2, last_row + 1)):
            # Check if the cells in columns B and C match the order number
            ajc_order_num = sheet.cell(row=row, column=2).value
            ajc_order_num = '' if ajc_order_num == None else str(ajc_order_num).lower().replace(" ", "")
            ctm_order_num = sheet.cell(row=row, column=3).value
            ctm_order_num = '' if ctm_order_num == None else str(ctm_order_num).lower().replace(" ", "")

            if ajc_order_num == orderNumber or ctm_order_num == orderNumber:
                customer = sheet_name
                ajc_order_number = sheet.cell(row=row, column=2).value
                ctm_order_number = sheet.cell(row=row, column=3).value
                # Display the columns F, L, and P if there is a match
                # print(f"Sheet: {sheet_name}, Row: {row}")
                item = sheet.cell(row=row, column=6).value
                
                order_qty = sheet.cell(row=row, column=12).value
                order_qty = '' if order_qty == None else order_qty

                prod_lb_col = 17 if customer == 'HEB' else 16
                load_date_col = 18 if customer == 'HEB' else 17
                
                prod_lb = sheet.cell(row=row, column=prod_lb_col).value
                prod_lb = 0 if prod_lb == None else prod_lb

                load_date = sheet.cell(row=row, column=load_date_col).value

                prod_qty = get_prod_qty(sheet.cell(row=row, column=10).value, prod_lb)

                output_row = {
                    "item": item, 
                    "order_qty": order_qty, 
                    "prod_qty": prod_qty,
                    "prod_lb": prod_lb,
                }
                output_table.insert(0, output_row)

    return {
        "ajc_order_number": ajc_order_number,
        "ctm_order_number": ctm_order_number,
        "customer": customer,
        "load_date": load_date,
        "table": output_table
    }


def search_in_petco_folder(order_num):
    current_year = datetime.datetime.now().year
    path_of_petco = f'../../PETCO发货/{current_year}'
    file = search_for_order_file(order_num, path_of_petco)


    customer = 'PETCO'
    load_date = None
    ajc_order_number = 'W0'
    ctm_order_number = order_num
    if file == None:
        output_table = []
    else:
        output_table = getPetcoOrderTable(file)

    return {
        "ajc_order_number": ajc_order_number,
        "ctm_order_number": ctm_order_number,
        "customer": customer,
        "load_date": load_date,
        "table": output_table
    }


def getPetcoOrderTable(file):
    # Open the Excel file and select the first worksheet
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    # Initialize an empty list to hold the orders
    orders = []
    
    # Loop through the rows starting from row 16
    for row in range(16, ws.max_row + 1):
        # Check if the value in column A is empty or 'Total:'
        if not ws.cell(row, 1).value or ws.cell(row, 1).value == 'Total:':
            break
        
        # Create a new order object and populate its properties
        order = {}
        order['order_qty'] = ws.cell(row, 1).value
        order['prod_qty'] = ws.cell(row, 1).value
        order['item'] = ws.cell(row, 7).value
        # order['product'] = ws.cell(row, 3).value
        # order['remark'] = ws.cell(row, 8).value
        
        # Append the order to the list of orders
        orders.append(order)
    
    # Return the list of orders
    return orders


def search_for_order_file(order_num, path_of_petco):
    for filename in os.listdir(path_of_petco):
        if order_num in filename:
            return os.path.join(path_of_petco, filename)
    return None


def get_prod_qty(string_param, prod_lb):
    if prod_lb == '':
        return ''
    unit = ''
    num = 0
    # Find the integer before 'OZ' using a regular expression
    match = re.search(r'(\d+(?:\.\d+)?)\s*[Oo][Zz]', string_param, flags=re.ASCII)

    # If a match is found, return the integer as an integer type
    if match:
        num = float(match.group(1))
        unit = 'oz'

    # Otherwise, try to find the integer before 'LB' using a regular expression
    else:
        match = re.search(r'(\d+(?:\.\d+)?)\s*[Ll][Bb]', string_param, flags=re.ASCII)

        # If a match is found, return the integer as an integer type, otherwise return None
        if match:
            num = float(match.group(1))
            unit = 'lb'
        else:
            return None

    if unit == 'oz':
        return prod_lb * 16 / num
    elif unit == 'lb':
        return prod_lb / num


def generateShippingList(order, petco_order):
    # print(order)
    try:
        customer_name = getCustomerName(order["customer"])
    except Exception as e:
        return 'Cannot find customer name'

    new_order = order.copy()

    try:
        if petco_order:
            new_order['table'] = process_orders(new_order['table'])
        else:
            new_order['table'] = process_orders(new_order['table'])

    except Exception as e:
        print(e)
        return 'An item in this order cannot be found'

    return update_template(customer_name, new_order, petco_order)


def getCustomerName(customer):
    # Open the workbook
    workbook = openpyxl.load_workbook('data.xlsx')

    # Select the Customer sheet
    sheet = workbook['Customers']

    # Iterate over each row in column A to find the search string
    for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if row[0] == customer:
            # Return the corresponding customer name in column B
            return row[1]

    # If the search string is not found, return None
    return 'CUSTOMER NAME NOT FOUND'


def process_orders(orders):
    # Load the Excel file
    wb = openpyxl.load_workbook(filename='data.xlsx', read_only=True)

    # Select the sheet named 'Items'
    ws = wb['Items']

    # Make a copy of the orders list
    new_orders = orders.copy()

    # Loop through the orders
    for order in new_orders:
        # Extract the numeric part of the item code
        item_num = int(re.findall('\d+', order['item'])[0])

        # Search for a match in the sheet
        found_match = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if int(re.findall('\d+', row[0])[0]) == item_num:
                # If a match is found, add the extra information to the order
                found_match = True
                order['item_num'] = row[1]
                order['case_cap'] = row[2]
                order['weight'] = row[3]
                order['remark'] = row[4]
                break

        if not found_match:
            raise f'{order["item"]} cannot be found in the data.xlsx, please add this item before proceed.'

    # Close the Excel file
    wb.close()

    return new_orders


# def process_petco_order(order):
#     print(order)


def update_template(customer_name, order, petco_order):
    final_message = ''
    wb = openpyxl.load_workbook('data.xlsx')
    sheet = wb['Template']
    
    # Append customer and order number to cells
    sheet['A5'] = f'TO: {customer_name}'
    sheet['G6'] = f'Ref No. {order["ajc_order_number"]}'
    
    # Find the next business day
    today = datetime.date.today()
    if today.weekday() == 4: # Friday
        next_biz_day = today + datetime.timedelta(days=3)
    else:
        next_biz_day = today + datetime.timedelta(days=1)
    sheet['G5'] = f"DATE: {next_biz_day.strftime('%m/%d/%Y')}" if order["load_date"] == None else f"DATE: {order['load_date'].strftime('%m/%d/%Y')}"

    cell_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
        right=openpyxl.styles.Side(style='thin'), 
        top=openpyxl.styles.Side(style='thin'), 
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # Insert rows for each item
    row_idx = 8
    for item in order['table']:
        if not petco_order and not weightVerification(item):
            final_message = 'The calculation is not correct, please create this shipping list manually.'

        sheet.insert_rows(row_idx, 1)
        sheet.row_dimensions[row_idx].height = 32
        sheet.cell(row=row_idx, column=1, value=order['ctm_order_number']).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=1).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=1).border = cell_border

        sheet.cell(row=row_idx, column=2, value=item['item']).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=2).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=2).border = cell_border

        sheet.cell(row=row_idx, column=3, value=item['item_num']).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=3).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=3).border = cell_border

        sheet.cell(row=row_idx, column=4, value=item['prod_qty']).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=4).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=4).border = cell_border

        sheet.cell(row=row_idx, column=5, value='=D{} / {}'.format(row_idx, item['case_cap'])).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=5).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=5).border = cell_border

        sheet.cell(row=row_idx, column=6, value='=D{} * {}'.format(row_idx, item['weight'])).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=6).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=6).border = cell_border

        sheet.cell(row=row_idx, column=7).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=7).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=7).border = cell_border

        sheet.cell(row=row_idx, column=8, value=item['remark']).font = openpyxl.styles.fonts.Font(name='Verdana')
        sheet.cell(row=row_idx, column=8).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
        sheet.cell(row=row_idx, column=8).border = cell_border

        row_idx += 1

    sheet.row_dimensions[row_idx].height = 32
    sheet.row_dimensions[row_idx + 1].height = 32

    sheet.cell(row=row_idx + 1, column=4, value=f"=SUM(D8:D{row_idx})")
    sheet.cell(row=row_idx + 1, column=5, value=f"=SUM(E8:E{row_idx})")
    sheet.cell(row=row_idx + 1, column=6, value=f"=SUM(F8:F{row_idx})")
    sheet.cell(row=row_idx + 1, column=7, value=f"=SUM(G8:G{row_idx})")

    sheet.row_dimensions[row_idx + 2].height = 31
    sheet.row_dimensions[row_idx + 3].height = 27
    sheet.row_dimensions[row_idx + 4].height = 27
    sheet.row_dimensions[row_idx + 5].height = 27
    sheet.row_dimensions[row_idx + 7].height = 15

    del wb['Customers']
    del wb['Items']

    wb['Template'].title = order["customer"]

    new_ctm_order_num = str(order["ctm_order_number"]).replace("/", "-")
        
    wb.save(f'./Shipping Lists/SHIPPING LIST-{order["ajc_order_number"]}-{new_ctm_order_num}.xlsx')
    wb.close()
    return f'SHIPPING LIST-{order["ajc_order_number"]}-{new_ctm_order_num}.xlsx is created.\n{final_message }'


def weightVerification(item):
    return True if item['prod_qty'] * item['weight'] == item['prod_lb'] else False


# Create the GUI
root = tk.Tk()
root.title("Shipping List Generator")
root.geometry("400x200") # Set the window size to 400 x 150 pixels
root.resizable(False, False) # Set the window to be non-resizable

# Add a text input field for the order number
orderEntry = tk.Entry(root, width=50, justify='center')
orderEntry.config(font=("Courier", 20)) # Increase the font size for better visibility
orderEntry.pack(side=tk.TOP, pady=10)
orderEntry.focus()

shippingListLabel = tk.Label(root, wraplength=280, justify='center', text="Enter an order number")
shippingListLabel.config(font=("Courier", 8)) # Increase the font size for better visibility
shippingListLabel.pack(side=tk.TOP)

# Add a button to generate the shipping list
generateButton = tk.Button(root, text="Generate Shipping List", command=onGenerateListButton, width=50, height=5, font=("Courier", 16))
generateButton.pack(side=tk.BOTTOM, pady=10)

# Start the main event loop
root.mainloop()